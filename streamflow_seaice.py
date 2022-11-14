import numpy as np
import pandas as pd
from scipy.signal import detrend
from scipy.stats import pearsonr, normaltest

extent = pd.read_excel('Sea_Ice_Index_Daily_Extent_G02135_v3.0.xlsx', sheet_name=1)
extent = extent.T
extent = extent.drop(['Unnamed: 0', 'Unnamed: 1', 1978, 2021, 2022, '1981-2010 mean', '1981-2010 median'])
extent = extent.astype('float')
extent = extent.interpolate(method='linear', axis=1)
extent = extent.interpolate(method='linear', axis=0)
extent_1d = extent.values.flatten()
seaice_detrended = detrend(extent_1d)
seaice = np.array(seaice_detrended).reshape(42, 366)
seaice = pd.DataFrame(seaice)


new_index = np.arange(41)

# get sea ice data for DJF
seaice_jf = seaice.loc[1:, 0:59]
seaice_jf.index = new_index
seaice_d = seaice.loc[0:40, 335:]
seaice_d.index = new_index
seaice_djf = pd.concat([seaice_d, seaice_jf], ignore_index=True, axis=1)
seaice_mean_djf = seaice_djf.mean(axis=1).to_list()

# get sea ice data for MAM
seaice_mam = seaice.loc[1:, 59:152]
seaice_mam = pd.DataFrame(seaice_mam)
seaice_mean_mam = seaice_mam.mean(axis=1).to_list()

# get sea ice data for JJA
seaice_jja = seaice.loc[1:, 152:243]
seaice_jja = pd.DataFrame(seaice_jja)
seaice_mean_jja = seaice_jja.mean(axis=1).to_list()

# oni = pd.read_excel('oni.xlsx')
# oni_djf = oni['DJF'].to_list()
# oni_mam = oni['MAM'].to_list()
# oni_jja = oni['JJA'].to_list()

num_stations = 131
seaice_corr = []
seaice_p = []

seasons = [seaice_mean_djf, seaice_mean_mam, seaice_mean_jja]

for season in range(len(seasons)):

    for sheet in range(num_stations):
        streamflow_df = pd.read_excel('Streamflow_Data.xlsx', sheet_name=sheet)
        # extract column with streamflow data
        streamflow_list = streamflow_df['ft3/s'].to_list()
        # convert data from ft^3/s to m^3/s
        streamflow_list = [value / 35.3147 for value in streamflow_list]
        # fill 0's in data to avoid RunTime Warning
        streamflow_list = [10**-10 if value == 0 else value for value in streamflow_list]
        # compute log_transformed streamflow
        streamfow_list = np.log10(streamflow_list)
        # compute mean and SD
        mean = np.mean(streamflow_list)
        sd = np.std(streamflow_list)
        # standardize the data
        standardized_values = []
        for item in range(len(streamflow_list)):
            standardized_values.append(((streamflow_list[item] - mean) / sd))

        june_streamflow = standardized_values[5::12]
        july_streamflow = standardized_values[6::12]
        august_streamflow = standardized_values[7::12]

        sf_df = pd.DataFrame([june_streamflow, july_streamflow, august_streamflow])
        sf_df = sf_df.T
        sf_df = sf_df.mean(axis=1).to_list()

        r = pearsonr(seasons[season], sf_df)

        seaice_corr.append(r[0])
        seaice_p.append(r[1])

corr_djf = seaice_corr[0:131]
corr_mam = seaice_corr[131:262]
corr_jja = seaice_corr[262:]

p_djf = seaice_p[0:131]
p_mam = seaice_p[131:262]
p_jja = seaice_p[262:]

df = []
for sheet in range(num_stations):
    print(sheet)
    streamflow_df = pd.read_excel('Streamflow_Data.xlsx', sheet_name=sheet)
    streamflow_subset = streamflow_df.drop(['Year', 'Month', 'ft3/s'], axis=1)
    streamflow_subset['r_djf'] = corr_djf[sheet]
    streamflow_subset['p_djf'] = p_djf[sheet]
    streamflow_subset['r_mam'] = corr_mam[sheet]
    streamflow_subset['p_mam'] = p_mam[sheet]
    streamflow_subset['r_jja'] = corr_jja[sheet]
    streamflow_subset['p_jja'] = p_jja[sheet]
    streamflow_subset = streamflow_subset.iloc[0]
    df.append(streamflow_subset)

from openpyxl import load_workbook

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True)
    return wb.sheetnames

sheet_names = get_sheetnames_xlsx('Streamflow_Data.xlsx')

df = pd.DataFrame(df)
df.index = sheet_names
df.to_excel('stats.xlsx')
